"""
Parse qudurat_responses.xlsx into structured JSON artifacts:
- question_index.json: answer keys + section/sub-section/skill/difficulty per question
- students.json: per-student answers + computed correctness + score summary

Run from the repo root:
    python3 data/process.py
"""
import json
import unicodedata
from pathlib import Path
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
XLSX = HERE / "qudurat_responses.xlsx"
OUT_INDEX = HERE / "question_index.json"
OUT_STUDENTS = HERE / "students.json"


def normalize_ar(s):
    """Normalize Arabic for robust equality.

    - Strip leading/trailing whitespace and collapse inner whitespace.
    - Remove tashkeel (combining diacritics).
    - Unify alef variants (أ إ آ → ا).
    - Unify ya variants (ى ئ → ي).
    - Unify ta marbuta (ة → ه).
    Handles typos in the answer key like المعطات vs المعطيات gracefully when
    both the key and the student answer collapse to the same normalized form.
    """
    if s is None:
        return ""
    s = str(s).strip()
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    s = s.replace("ى", "ي").replace("ئ", "ي")
    s = s.replace("ة", "ه")
    return " ".join(s.split())


def main():
    wb = load_workbook(XLSX, data_only=True)
    idx = wb["Index"]

    questions = {"quant": {}, "verbal": {}}
    for r in range(5, 55):
        # Quant
        if idx.cell(r, 2).value is not None:
            qn = int(idx.cell(r, 2).value)
            questions["quant"][qn] = {
                "qn": qn,
                "correct_answer": idx.cell(r, 3).value,
                "section": "كمي",
                "sub_section": idx.cell(r, 9).value,
                "skill": idx.cell(r, 10).value,
                "difficulty": int(idx.cell(r, 11).value),
            }
        # Verbal (answer key in col 6; mapping in cols 14 (5-cat skill) / 15 (3-cat sub-section) / 16 (difficulty))
        if idx.cell(r, 5).value is not None:
            qn = int(idx.cell(r, 5).value)
            questions["verbal"][qn] = {
                "qn": qn,
                "correct_answer": idx.cell(r, 6).value,
                "section": "لفظي",
                "sub_section": idx.cell(r, 15).value,  # 3-category
                "skill": idx.cell(r, 14).value,  # 5-category
                "difficulty": int(idx.cell(r, 16).value),
            }

    with OUT_INDEX.open("w", encoding="utf-8") as f:
        json.dump(questions, f, ensure_ascii=False, indent=2)

    resp = wb["Responses"]
    students = []
    for row in range(2, resp.max_row + 1):
        name = resp.cell(row, 3).value
        if not name:
            continue
        phone = resp.cell(row, 6).value
        meta = {
            "timestamp": str(resp.cell(row, 1).value) if resp.cell(row, 1).value else "",
            "name": (name or "").strip(),
            "email": (resp.cell(row, 4).value or "").strip(),
            "gender": (resp.cell(row, 5).value or "").strip(),
            "student_phone": str(phone).strip() if phone else "",
            "parent_phone": str(resp.cell(row, 7).value).strip() if resp.cell(row, 7).value else "",
            "school": (resp.cell(row, 8).value or "").strip(),
            "grade": (resp.cell(row, 9).value or "").strip(),
        }
        # PIN: last 4 digits of student phone
        digits = "".join(c for c in meta["student_phone"] if c.isdigit())
        meta["pin"] = digits[-4:] if len(digits) >= 4 else "0000"

        quant_answers = []
        for q in range(1, 51):
            ans = resp.cell(row, 10 + q).value
            key = questions["quant"][q]["correct_answer"]
            correct = normalize_ar(ans) == normalize_ar(key) and ans is not None
            quant_answers.append({
                "qn": q,
                "student_answer": ans,
                "correct": correct,
                "sub_section": questions["quant"][q]["sub_section"],
                "skill": questions["quant"][q]["skill"],
                "difficulty": questions["quant"][q]["difficulty"],
            })

        verbal_answers = []
        for q in range(1, 51):
            ans = resp.cell(row, 60 + q).value
            key = questions["verbal"][q]["correct_answer"]
            correct = normalize_ar(ans) == normalize_ar(key) and ans is not None
            verbal_answers.append({
                "qn": q,
                "student_answer": ans,
                "correct": correct,
                "sub_section": questions["verbal"][q]["sub_section"],
                "skill": questions["verbal"][q]["skill"],
                "difficulty": questions["verbal"][q]["difficulty"],
            })

        quant_score = sum(1 for a in quant_answers if a["correct"])
        verbal_score = sum(1 for a in verbal_answers if a["correct"])
        students.append({
            **meta,
            "quant_score": quant_score,
            "verbal_score": verbal_score,
            "total_score": quant_score + verbal_score,
            "overall_pct": round((quant_score + verbal_score) / 100 * 100),
            "quant_answers": quant_answers,
            "verbal_answers": verbal_answers,
        })

    with OUT_STUDENTS.open("w", encoding="utf-8") as f:
        json.dump(students, f, ensure_ascii=False, indent=2)

    print(f"Wrote {OUT_INDEX} ({len(questions['quant'])} quant + {len(questions['verbal'])} verbal questions)")
    print(f"Wrote {OUT_STUDENTS} ({len(students)} students)")
    print("\nScores summary:")
    for s in students:
        print(f"  {s['name']}: Q {s['quant_score']}/50  V {s['verbal_score']}/50  Total {s['total_score']}/100 ({s['overall_pct']}%)  PIN {s['pin']}")


if __name__ == "__main__":
    main()
