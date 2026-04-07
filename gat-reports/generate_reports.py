#!/usr/bin/env python3
"""
GAT Report Generator
====================
Reads GAT_Question_Bank.xlsx + a response CSV, generates one HTML report per student.

Usage:
    python generate_reports.py \
        --question-bank ../GAT_Question_Bank.xlsx \
        --responses ../GAT\ Sample\ Response\ v1.csv \
        --output-dir docs/ \
        --logo-path ../noon\ logo\ v2.png

Output:
    docs/<student-slug>.html   (one per student)
    docs/index.html            (directory listing, optional)
"""

import argparse
import base64
import csv
import json
import math
import os
import re
from datetime import datetime

import openpyxl


# ── Question classification ──────────────────────────────────────────────────

QUANT_CATEGORIES = {
    "Algebra": [1,4,12,13,18,20,24,27,29,31,32,33,35,37,38,39,40,42,44,45],
    "Arithmetic & Number Sense": [2,6,8,11,21,30,36],
    "Ratios, Percentages & Proportions": [3,5,7,9,10,15,16,17,26],
    "Comparison & Estimation": [19,23,25,28,34,41,46,54],
    "Geometry": [14,43,47,48,49,50,51,52,53],
    "Data & Statistics": [22,55,56,57,58,59,60],
}

QUANT_NUM_TO_CAT = {}
for cat, nums in QUANT_CATEGORIES.items():
    for n in nums:
        QUANT_NUM_TO_CAT[n] = cat

VERBAL_MICRO = {
    1:"Antonyms",2:"Synonyms",3:"Object-Property",4:"Part-Whole",5:"Association",
    6:"Part-Whole",7:"Tool-Function",8:"Degree/Intensity",9:"Antonyms",10:"Antonyms",
    11:"Category Recognition",12:"Category Recognition",13:"Category Recognition",
    14:"Category Recognition",15:"Category Recognition",16:"Category Recognition",
    17:"Category Recognition",18:"Category Recognition",19:"Category Recognition",20:"Category Recognition",
    21:"Advanced Vocabulary",22:"Advanced Vocabulary",23:"Advanced Vocabulary",
    24:"Contextual Meaning",25:"Word Usage",26:"Contradiction Detection",
    27:"Contradiction Detection",28:"Contradiction Detection",29:"Contradiction Detection",30:"Contradiction Detection",
    31:"Context Clues",32:"Context Clues",33:"Context Clues",34:"Context Clues",
    35:"Context Clues",36:"Prepositions & Idioms",37:"Verb Tense & Usage",
    38:"Vocabulary in Context",39:"Vocabulary in Context",40:"Vocabulary in Context",
    41:"Pronoun Reference",42:"Inference",43:"Inference",44:"Main Idea & Theme",
    45:"Main Idea & Theme",46:"Detail Extraction",47:"Detail Extraction",
    48:"Detail Extraction",49:"Detail Extraction",50:"Detail Extraction",
}

QUANT_MICRO = {
    1:"Algebraic Fractions",2:"Place Value & Number Sense",3:"Inverse Proportion",
    4:"Solving Cubic Equations",5:"Percentage Calculation",6:"Number Line & Integers",
    7:"Speed, Distance, Time",8:"Arithmetic Sequences",9:"Fractions & Square Roots",
    10:"Percentage Calculation",11:"Pattern Recognition (Fibonacci)",12:"Algebraic Comparison",
    13:"Linear Equations",14:"Angle Properties",15:"Profit & Loss (Percentages)",
    16:"Percentage Decrease",17:"Proportional Reasoning",18:"Linear Equations",
    19:"Absolute Value & Estimation",20:"Simplifying Radicals",21:"Order of Operations (BODMAS)",
    22:"Averages (Mean)",23:"Algebraic Inequalities",24:"Exponent Rules",
    25:"Mean & Exponents",26:"Relative Speed",27:"Linear Equations",
    28:"Properties of Negative Cubes",29:"Solving Radical Equations",30:"Exponent Rules",
    31:"Linear Expressions (Factoring)",32:"Difference of Squares",33:"Nested Radicals",
    34:"Square Roots & Arithmetic",35:"Linear Inequalities",36:"Summation & Series",
    37:"Substitution & Cubes",38:"Rational Equations",39:"Exponent Equations",
    40:"Polynomial Expansion (FOIL)",41:"Algebraic Comparison",42:"Translating Words to Algebra",
    43:"Area of Squares",44:"Algebraic Identities",45:"Exponent Simplification",
    46:"Comparison with Fractions",47:"Circle Circumference",48:"Inscribed Angles",
    49:"Triangle Angle Sum",50:"Area of Rectangles",51:"Supplementary Angles",
    52:"Supplementary Angles",53:"Complementary Angles",54:"Algebraic Comparison",
    55:"Averages (Mean)",56:"Venn Diagrams & Probability",57:"Reading Data from Charts",
    58:"Permutations",59:"Pie Chart Interpretation",60:"Range, Median & Number Theory",
}

STUDY_GROUPS = {
    "Angles & Angle Relationships": ["Supplementary Angles","Complementary Angles","Angle Properties","Inscribed Angles"],
    "Geometry Formulas": ["Circle Circumference","Area of Squares","Area of Rectangles","Triangle Angle Sum"],
    "Algebraic Expressions": ["Algebraic Identities","Difference of Squares","Polynomial Expansion (FOIL)","Linear Expressions (Factoring)"],
    "Number Operations": ["Order of Operations (BODMAS)","Simplifying Radicals","Square Roots & Arithmetic","Substitution & Cubes"],
    "Word Relationships": ["Antonyms","Synonyms"],
    "Reading for Meaning": ["Pronoun Reference","Inference"],
    "Rates & Proportions": ["Inverse Proportion","Speed, Distance, Time","Relative Speed"],
    "Sequences & Summation": ["Summation & Series","Arithmetic Sequences","Pattern Recognition (Fibonacci)"],
    "Data & Charts": ["Reading Data from Charts","Permutations","Range, Median & Number Theory","Averages (Mean)"],
    "Comparison Strategies": ["Algebraic Comparison","Absolute Value & Estimation","Mean & Exponents","Comparison with Fractions"],
    "Reading Comprehension": ["Detail Extraction","Main Idea & Theme"],
    "Vocabulary & Context": ["Vocabulary in Context","Context Clues","Prepositions & Idioms","Verb Tense & Usage"],
    "Linear Equations": ["Linear Equations","Linear Inequalities","Algebraic Fractions"],
    "Exponents & Radicals": ["Exponent Rules","Exponent Equations","Nested Radicals","Solving Radical Equations","Exponent Simplification"],
    "Percentage & Proportion Problems": ["Percentage Calculation","Percentage Decrease","Profit & Loss (Percentages)","Proportional Reasoning"],
}

STUDY_GROUP_DESCRIPTIONS = {
    "Angles & Angle Relationships": "Supplementary, complementary, and inscribed angles",
    "Geometry Formulas": "Circumference, area of squares, rectangles, triangle angles",
    "Algebraic Expressions": "Factoring, expanding, difference of squares, identities",
    "Number Operations": "Order of operations, radicals, roots, substitution",
    "Word Relationships": "Synonyms, antonyms, and degree relationships",
    "Reading for Meaning": "Pronoun reference and drawing inferences",
    "Rates & Proportions": "Inverse proportion, speed/distance/time, relative speed",
    "Sequences & Summation": "Arithmetic sequences, Fibonacci, summation formulas",
    "Data & Charts": "Reading charts, averages, permutations, range & median",
    "Comparison Strategies": "Algebraic comparison, absolute value, estimation",
    "Reading Comprehension": "Locating and extracting specific facts from passages",
    "Vocabulary & Context": "Determining word meaning from surrounding text",
    "Linear Equations": "Solving for x, inequalities, algebraic fractions",
    "Exponents & Radicals": "Exponent rules, equations, nested and simple radicals",
    "Percentage & Proportion Problems": "Percentage increase/decrease, profit & loss, proportional reasoning",
}


# ── Data loading ─────────────────────────────────────────────────────────────

def load_question_bank(path):
    wb = openpyxl.load_workbook(path)

    verbal_questions = []
    for i, row in enumerate(wb['Verbal'].iter_rows(min_row=2, values_only=True)):
        verbal_questions.append({
            'num': i+1, 'type': row[1],
            'question': str(row[3]) if row[3] else '',
            'difficulty': row[9],
        })

    quant_questions = []
    for i, row in enumerate(wb['Quantitative'].iter_rows(min_row=2, values_only=True)):
        quant_questions.append({
            'num': i+1,
            'question': str(row[1]) if row[1] else '',
            'difficulty': row[9],
            'type': QUANT_NUM_TO_CAT.get(i+1, 'Unknown'),
        })

    return verbal_questions, quant_questions


def parse_score(score_str):
    try:
        return float(score_str.split('/')[0].strip()) >= 1.0
    except:
        return False


def map_quant_difficulty(d):
    d = float(d)
    if d <= 2: return 'Easy'
    if d == 3: return 'Medium'
    return 'Hard'


def load_responses(path, verbal_questions, quant_questions):
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers = next(reader)
        rows = list(reader)

    # Find question columns
    q_cols = []
    for i, h in enumerate(headers):
        if '[Score]' not in h and '[Feedback]' not in h:
            m = re.match(r'^(\d+)\.\s*(.*)', h)
            if m:
                q_cols.append((i, int(m.group(1)), h))

    # Split verbal/quant at numbering reset
    prev, split_idx = 0, len(q_cols)
    for j, (idx, qnum, h) in enumerate(q_cols):
        if qnum < prev:
            split_idx = j
            break
        prev = qnum

    verbal_cols = q_cols[:split_idx]
    quant_cols = q_cols[split_idx:]

    students = []
    for row in rows:
        name = row[2].strip()
        if not name or name.lower() == 'test':
            continue

        grade = row[8].strip()
        timestamp = row[0].strip()

        # Parse date
        try:
            dt = datetime.strptime(timestamp.split(' GMT')[0].strip(), '%Y/%m/%d %I:%M:%S %p')
            date_str = dt.strftime('%B %-d, %Y')
        except:
            date_str = timestamp.split(' ')[0] if timestamp else ''

        # Phone number (last 4 digits for PIN)
        phone = row[14].strip() if len(row) > 14 else ''
        pin = phone[-4:] if len(phone) >= 4 else '0000'

        # Parse verbal
        verbal_results = []
        for j, (col_idx, qnum, header) in enumerate(verbal_cols):
            qinfo = verbal_questions[j] if j < len(verbal_questions) else {}
            is_correct = parse_score(row[col_idx + 1])
            micro = VERBAL_MICRO.get(qnum, 'Unknown')
            verbal_results.append({
                'qnum': qnum,
                'skill': qinfo.get('type', 'Unknown'),
                'difficulty': qinfo.get('difficulty', 'Unknown'),
                'correct': is_correct,
                'micro_skill': micro,
            })

        # Parse quant
        quant_results = []
        for j, (col_idx, qnum, header) in enumerate(quant_cols):
            qinfo = quant_questions[j] if j < len(quant_questions) else {}
            is_correct = parse_score(row[col_idx + 1])
            micro = QUANT_MICRO.get(qnum, 'Unknown')
            raw_diff = qinfo.get('difficulty', 1)
            quant_results.append({
                'qnum': qnum,
                'skill': qinfo.get('type', 'Unknown'),
                'difficulty': map_quant_difficulty(raw_diff),
                'correct': is_correct,
                'micro_skill': micro,
            })

        students.append({
            'name': name,
            'first_name': name.split()[0],
            'grade': grade,
            'date_str': date_str,
            'pin': pin,
            'verbal_results': verbal_results,
            'quant_results': quant_results,
        })

    return students


# ── Analytics ────────────────────────────────────────────────────────────────

def aggregate_by_skill(results):
    skills = {}
    for r in results:
        sk = r['skill']
        if sk not in skills:
            skills[sk] = {'total': 0, 'correct': 0, 'by_difficulty': {}}
        skills[sk]['total'] += 1
        if r['correct']:
            skills[sk]['correct'] += 1
        d = r['difficulty']
        if d not in skills[sk]['by_difficulty']:
            skills[sk]['by_difficulty'][d] = {'total': 0, 'correct': 0}
        skills[sk]['by_difficulty'][d]['total'] += 1
        if r['correct']:
            skills[sk]['by_difficulty'][d]['correct'] += 1
    return skills


def difficulty_summary(results, levels):
    summary = []
    for lbl in levels:
        total = sum(1 for r in results if r['difficulty'] == lbl)
        correct = sum(1 for r in results if r['difficulty'] == lbl and r['correct'])
        if total > 0:
            summary.append((lbl, correct, total, round(correct/total*100)))
    return summary


def compute_study_groups(verbal_results, quant_results):
    all_results = verbal_results + quant_results
    micro_counts = {}
    for r in all_results:
        ms = r['micro_skill']
        if ms not in micro_counts:
            micro_counts[ms] = {'correct': 0, 'total': 0}
        micro_counts[ms]['total'] += 1
        if r['correct']:
            micro_counts[ms]['correct'] += 1

    groups = []
    for group_name, skills in STUDY_GROUPS.items():
        total = sum(micro_counts.get(s, {}).get('total', 0) for s in skills)
        correct = sum(micro_counts.get(s, {}).get('correct', 0) for s in skills)
        if total > 0:
            pct = round(correct / total * 100)
            wrong = total - correct
            groups.append((group_name, correct, total, pct, wrong))

    # Top 3 nailed (highest pct, tiebreak by sample size)
    nailed = sorted([g for g in groups if g[3] >= 75], key=lambda x: (-x[3], -x[2]))[:3]
    # Top 3 to practice (most wrong answers, tiebreak by lowest pct)
    practice = sorted([g for g in groups if g[3] < 75], key=lambda x: (-x[4], x[3]))[:3]

    return nailed, practice


# ── Score tier ───────────────────────────────────────────────────────────────

READINESS_TIERS = [
    {"label": "Foundation",  "sub": "Fix knowledge gaps",      "min": 0,  "max": 49,  "color": "#EA8CBB"},
    {"label": "Development", "sub": "Build & practice skills",  "min": 50, "max": 89,  "color": "#5FA0E0"},
    {"label": "Advanced",    "sub": "Targeted mastery",         "min": 90, "max": 100, "color": "#4BA283"},
]


def score_tier(pct):
    if pct >= 90: return ("Advanced Phase", "#4BA283")
    if pct >= 50: return ("Development Phase", "#5FA0E0")
    return ("Foundation Phase", "#EA8CBB")


def readiness_tier_index(pct):
    for i in range(len(READINESS_TIERS) - 1, -1, -1):
        if pct >= READINESS_TIERS[i]["min"]:
            return i
    return 0


def readiness_bar_html(overall_pct):
    idx = readiness_tier_index(overall_pct)
    current = READINESS_TIERS[idx]
    # Bar segments
    segs = ''
    for i, t in enumerate(READINESS_TIERS):
        w = t["max"] - t["min"] + 1
        op = 1 if i <= idx else 0.15
        div = '<div class="seg-divider"></div>' if i > 0 else ''
        segs += f'{div}<div class="seg" style="flex:{w};background:{t["color"]};opacity:{op}"></div>'
    # Marker position
    marker_left = max(8, min(92, overall_pct))
    marker_label = f'{current["label"]} Phase &middot; {overall_pct}%'
    # Labels
    labels = ''
    for i, t in enumerate(READINESS_TIERS):
        w = t["max"] - t["min"] + 1
        cls = 'seg-label'
        if i == idx: cls += ' active'
        elif i > idx: cls += ' future'
        rng = f'Below {t["max"]+1}%' if t["min"] == 0 else (f'{t["min"]}%+' if t["max"] == 100 else f'{t["min"]}&ndash;{t["max"]}%')
        labels += f'<div class="{cls}" style="flex:{w}"><div class="seg-label-name">{t["label"]}</div><div class="seg-label-sub">{t["sub"]}</div><div class="seg-label-range">{rng}</div></div>'
    return f'''<div class="readiness-scale">
      <div class="readiness-heading">Readiness Scale</div>
      <div class="seg-bar">{segs}</div>
      <div class="seg-marker-wrap"><div class="seg-marker" style="left:{marker_left}%"><div class="seg-marker-arrow"></div><div class="seg-marker-label">{marker_label}</div></div></div>
      <div class="seg-labels">{labels}</div>
    </div>'''


# ── HTML generation ──────────────────────────────────────────────────────────

def pct(correct, total):
    return round(correct / total * 100) if total > 0 else 0


def bar_class(p):
    if p >= 65: return 'bg'
    if p >= 35: return 'by'
    return 'br'


def color_var(p):
    if p >= 70: return '--dk-green'
    if p >= 40: return '--orange'
    return '--pink'


def tag_class(p):
    if p >= 70: return 'tg', 'Strong'
    if p >= 40: return 'td', 'Developing'
    return 'tw', 'Needs work'


def skill_card_html(name, score_pct, skill_data):
    tc, tl = tag_class(score_pct)
    cv = color_var(score_pct)
    bar_bg = "var(--green)" if score_pct >= 70 else "var(--orange)" if score_pct >= 40 else "var(--pink)"

    diff_items = ''
    for d in ['Easy', 'Medium', 'Hard']:
        if d in skill_data['by_difficulty']:
            dd = skill_data['by_difficulty'][d]
            dp = pct(dd['correct'], dd['total'])
            dc = color_var(dp)
            diff_items += f'<div class="sk-diff-item"><div class="dn">{d}</div><div class="dv" style="color:var({dc})">{dp}%</div><div class="dc">{dd["correct"]}/{dd["total"]}</div></div>'

    return f'''
        <div class="skill-card" onclick="this.classList.toggle('expanded')">
          <div class="sk-top">
            <div><div class="sk-name">{name}</div><span class="sk-tag {tc}">{tl}</span></div>
            <div class="sk-pct" style="color:var({cv})">{score_pct}<span class="sk-pct-u">%</span></div>
          </div>
          <div class="sk-bar"><div class="sk-bar-fill" style="width:{score_pct}%;background:{bar_bg}"></div></div>
          <div class="sk-detail"><div class="sk-diff-grid">{diff_items}</div></div>
        </div>'''


def micro_card_html(name, correct, total, score_pct, color, desc):
    return f'''
          <div class="micro-card">
            <div class="mc-name">{name}</div>
            <div class="mc-score" style="color:var({color})">{score_pct}<span style="font-size:14px">%</span></div>
            <div class="mc-raw">{correct} / {total} correct</div>
            <div class="mc-bar"><div class="mc-bar-fill" style="width:{score_pct}%;background:var({color.replace('dk-','')})"></div></div>
            <p style="font-size:11px;color:#aaa;margin-top:8px;line-height:1.4">{desc}</p>
          </div>'''


def generate_report(student, logo_b64):
    s = student
    v_results = s['verbal_results']
    q_results = s['quant_results']
    v_score = sum(1 for r in v_results if r['correct'])
    q_score = sum(1 for r in q_results if r['correct'])
    v_total = len(v_results)
    q_total = len(q_results)
    overall = v_score + q_score
    overall_total = v_total + q_total
    overall_pct = pct(overall, overall_total)
    verbal_pct = pct(v_score, v_total)
    quant_pct = pct(q_score, q_total)

    tier_label, tier_color = score_tier(overall_pct)

    # Ring
    circ = 459
    ring_offset = round(circ * (1 - overall_pct / 100))

    # Skills
    v_skills = aggregate_by_skill(v_results)
    q_skills = aggregate_by_skill(q_results)

    # Difficulty summaries
    v_diff = difficulty_summary(v_results, ['Easy', 'Medium', 'Hard'])
    q_diff = difficulty_summary(q_results, ['Easy', 'Medium', 'Hard'])

    # Strengths / focus
    all_skills = []
    for sk, d in v_skills.items():
        p = pct(d['correct'], d['total'])
        all_skills.append((sk, 'Verbal', p))
    for sk, d in q_skills.items():
        p = pct(d['correct'], d['total'])
        all_skills.append((sk, 'Quantitative', p))

    strengths = sorted([x for x in all_skills if x[2] >= 70], key=lambda x: -x[2])[:5]
    focus = sorted([x for x in all_skills if x[2] < 70], key=lambda x: x[2])[:5]

    # Micro-skill study groups
    nailed, practice = compute_study_groups(v_results, q_results)

    # Verbal skill cards (sorted high to low)
    v_sorted = sorted(v_skills.items(), key=lambda x: -pct(x[1]['correct'], x[1]['total']))
    q_sorted = sorted(q_skills.items(), key=lambda x: -pct(x[1]['correct'], x[1]['total']))

    # ── Build verbal insight
    v_strong_count = sum(1 for _, d in v_skills.items() if pct(d['correct'], d['total']) >= 70)
    v_top_skill = v_sorted[0][0] if v_sorted else 'N/A'
    v_top_pct = pct(v_sorted[0][1]['correct'], v_sorted[0][1]['total']) if v_sorted else 0
    v_weak_skill = v_sorted[-1][0] if v_sorted else 'N/A'
    v_weak_pct = pct(v_sorted[-1][1]['correct'], v_sorted[-1][1]['total']) if v_sorted else 0

    verbal_insight = (
        f"{s['first_name']}, your verbal reasoning {'is strong' if verbal_pct >= 70 else 'is developing' if verbal_pct >= 50 else 'has room to grow'}. "
        f"You scored above 70% in {v_strong_count} out of {len(v_skills)} skill areas. "
        f"{v_top_skill} is your top skill at {v_top_pct}%. "
        f"{v_weak_skill} is the area to focus on at {v_weak_pct}%. "
        f"Your Easy and Medium performance is consistent, and the main opportunity is in harder question patterns."
    )

    # ── Build quant insight
    q_strong_count = sum(1 for _, d in q_skills.items() if pct(d['correct'], d['total']) >= 70)
    q_top_skill = q_sorted[0][0] if q_sorted else 'N/A'
    q_top_pct = pct(q_sorted[0][1]['correct'], q_sorted[0][1]['total']) if q_sorted else 0
    q_weak_skill = q_sorted[-1][0] if q_sorted else 'N/A'
    q_weak_pct = pct(q_sorted[-1][1]['correct'], q_sorted[-1][1]['total']) if q_sorted else 0

    quant_insight = (
        f"{s['first_name']}, your quantitative score shows {q_top_skill} as {'a strength' if q_top_pct >= 70 else 'your strongest area'} at {q_top_pct}%. "
        f"{q_weak_skill} ({q_weak_pct}%) and the other lower-scoring areas are where the most points were left on the table. "
        f"You handle Easy questions at {q_diff[0][3] if q_diff else 0}%, and there's a drop into Medium ({q_diff[1][3] if len(q_diff) > 1 else 0}%) "
        f"and Hard ({q_diff[2][3] if len(q_diff) > 2 else 0}%) where targeted practice will make the biggest difference."
    )

    # ── Build overview snapshot
    snapshot = (
        f"{s['first_name']}, you scored <span class=\"hl\">{overall} out of {overall_total}</span> on the GAT diagnostic ({overall_pct}%). "
        f"Your verbal ability {'is a clear strength' if verbal_pct >= 70 else 'is developing'} at <span class=\"hl\">{verbal_pct}%</span>, "
        f"{'with ' + str(v_strong_count) + ' out of ' + str(len(v_skills)) + ' skill areas above 70%' if v_strong_count > 0 else 'with room to grow across skill areas'}. "
        f"Quantitative reasoning sits at <span class=\"hl\">{quant_pct}%</span>, with {q_top_skill} leading at {q_top_pct}% "
        f"and {q_weak_skill} as the main area to work on. "
        f"You handle Easy and Medium questions consistently well. The biggest opportunity is in harder questions, "
        f"where focused practice will have the most impact on your overall score."
    )

    # ── Strengths/focus HTML
    def dual_list(items, color):
        return ''.join(
            f'<li>{sk} <span class="sc" style="color:var({color})">{p}%</span><br><span class="sec">{sec}</span></li>'
            for sk, sec, p in items
        ) or f'<li style="color:#bbb">None identified</li>'

    # ── Skill cards HTML
    verbal_cards = ''.join(
        skill_card_html(sk, pct(d['correct'], d['total']), d)
        for sk, d in v_sorted
    )
    quant_cards = ''.join(
        skill_card_html(sk, pct(d['correct'], d['total']), d)
        for sk, d in q_sorted
    )

    # ── Micro-skill cards
    nailed_html = ''.join(
        micro_card_html(n, c, t, p, '--dk-green', STUDY_GROUP_DESCRIPTIONS.get(n, ''))
        for n, c, t, p, _ in nailed
    )
    practice_html = ''.join(
        micro_card_html(n, c, t, p, '--pink', STUDY_GROUP_DESCRIPTIONS.get(n, ''))
        for n, c, t, p, _ in practice
    )

    # ── Difficulty bars
    def diff_bars(summary):
        html = ''
        for lbl, c, t, p in summary:
            bc = bar_class(p)
            html += f'<div class="dr"><div class="dl">{lbl}</div><div class="dt"><div class="df {bc}" style="width:{p}%">{p}%</div></div><div class="ds">{c}/{t}</div></div>\n'
        return html

    grade_num = s['grade'].replace('G', '')

    # ── Full HTML
    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>GAT Diagnostic - {s['name']}</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<style>
:root {{
  --green:#64D8AE;--dk-green:#4BA283;--lt-green:#83E0BE;
  --black:#141414;--off-white:#F9F5EF;
  --purple:#B08AF9;--blue:#5FA0E0;--orange:#FF703E;--pink:#EA8CBB;
  --gray:#BBB8B3;--lt-gray:#E8E5E0;
}}
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',-apple-system,BlinkMacSystemFont,sans-serif;background:var(--off-white);color:var(--black);-webkit-font-smoothing:antialiased}}
.lock-screen{{min-height:100vh;display:flex;flex-direction:column;align-items:center;justify-content:center;background:white;padding:32px 24px;text-align:center;position:relative;overflow:hidden}}
.lock-screen::before{{content:'';position:absolute;top:-60px;right:-60px;width:220px;height:220px;background:radial-gradient(circle,rgba(100,216,174,0.15) 0%,transparent 70%);border-radius:50%}}
.lock-screen::after{{content:'';position:absolute;bottom:-40px;left:-40px;width:160px;height:160px;background:radial-gradient(circle,rgba(176,138,249,0.1) 0%,transparent 70%);border-radius:50%}}
.lock-logo{{width:64px;margin-bottom:32px;position:relative;z-index:1}}
.lock-title{{font-size:22px;font-weight:300;color:var(--black);margin-bottom:6px;position:relative;z-index:1}}
.lock-sub{{font-size:14px;color:var(--gray);margin-bottom:32px;position:relative;z-index:1}}
.pin-inputs{{display:flex;gap:12px;justify-content:center;margin-bottom:24px;position:relative;z-index:1}}
.pin-inputs input{{width:52px;height:60px;border:2px solid var(--lt-gray);border-radius:12px;text-align:center;font-size:24px;font-weight:300;font-family:'Inter',sans-serif;color:var(--black);outline:none;transition:border-color 0.2s;-moz-appearance:textfield}}
.pin-inputs input::-webkit-outer-spin-button,.pin-inputs input::-webkit-inner-spin-button{{-webkit-appearance:none;margin:0}}
.pin-inputs input:focus{{border-color:var(--green)}}
.pin-inputs input.error{{border-color:var(--pink);animation:shake 0.4s ease}}
@keyframes shake{{0%,100%{{transform:translateX(0)}}25%{{transform:translateX(-6px)}}75%{{transform:translateX(6px)}}}}
.lock-error{{font-size:13px;color:var(--pink);height:20px;position:relative;z-index:1}}
.lock-hidden{{display:none}}
.hero{{background:white;padding:32px 24px 60px;text-align:center;position:relative;overflow:hidden}}
.hero::before{{content:'';position:absolute;top:-80px;right:-80px;width:280px;height:280px;background:radial-gradient(circle,rgba(100,216,174,0.2) 0%,transparent 70%);border-radius:50%;pointer-events:none}}
.hero::after{{content:'';position:absolute;bottom:-60px;left:-40px;width:200px;height:200px;background:radial-gradient(circle,rgba(176,138,249,0.12) 0%,transparent 70%);border-radius:50%;pointer-events:none}}
.hero-logo{{width:72px;margin-bottom:24px;position:relative;z-index:1}}
.student-name{{font-size:38px;font-weight:300;letter-spacing:-0.8px;color:var(--black);margin-bottom:4px;line-height:1.15;position:relative;z-index:1;word-break:break-word}}
.student-meta{{font-size:13px;color:var(--gray);font-weight:500;margin-bottom:32px;position:relative;z-index:1;letter-spacing:0.3px}}
.score-ring-area{{display:flex;justify-content:center;margin-bottom:16px;position:relative;z-index:1}}
.score-ring{{position:relative;width:170px;height:170px}}
.score-ring svg{{transform:rotate(-90deg)}}
.ring-bg{{fill:none;stroke:var(--lt-gray);stroke-width:10}}
.ring-fill{{fill:none;stroke:var(--green);stroke-width:10;stroke-linecap:round;stroke-dasharray:{circ};stroke-dashoffset:{circ};animation:ring-in 1.2s ease forwards 0.3s}}
@keyframes ring-in{{to{{stroke-dashoffset:{ring_offset}}}}}
.ring-label{{position:absolute;inset:0;display:flex;flex-direction:column;align-items:center;justify-content:center}}
.ring-label .pct{{font-size:42px;font-weight:300;color:var(--black);line-height:1}}.ring-label .pct-unit{{font-size:18px;font-weight:300}}
.ring-label .raw{{font-size:12px;color:var(--gray);margin-top:4px;font-weight:500}}
.readiness-scale{{position:relative;z-index:1;padding:0 4px;margin-bottom:8px}}
.readiness-heading{{font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:1px;color:var(--gray);margin-bottom:12px;text-align:center}}
.seg-bar{{display:flex;height:12px;border-radius:7px;overflow:hidden;background:var(--lt-gray)}}
.seg{{height:100%;transition:opacity 0.4s}}.seg-divider{{width:2px;background:white;flex-shrink:0}}
.seg-marker-wrap{{position:relative;height:32px;margin-top:5px}}
.seg-marker{{position:absolute;transform:translateX(-50%);display:flex;flex-direction:column;align-items:center}}
.seg-marker-arrow{{width:0;height:0;border-left:5px solid transparent;border-right:5px solid transparent;border-bottom:6px solid var(--black)}}
.seg-marker-label{{background:var(--black);color:white;font-size:10px;font-weight:600;padding:4px 12px;border-radius:10px;white-space:nowrap}}
.seg-labels{{display:flex;margin-top:8px}}
.seg-label{{display:flex;flex-direction:column;align-items:center;text-align:center}}
.seg-label-name{{font-size:10px;font-weight:600;line-height:1.3;color:#666}}
.seg-label-sub{{font-size:9px;color:#777;margin-top:2px;line-height:1.3;font-style:italic}}
.seg-label-range{{font-size:9px;color:#888;margin-top:2px}}
.seg-label.active .seg-label-name{{color:var(--black);font-weight:700}}
.seg-label.active .seg-label-sub{{color:#444}}
.seg-label.active .seg-label-range{{color:#666}}
.seg-label.future .seg-label-name{{color:#999}}
.seg-label.future .seg-label-sub{{color:#aaa}}
.seg-label.future .seg-label-range{{color:#aaa}}
.score-nav{{display:grid;grid-template-columns:repeat(3,1fr);gap:12px;padding:0 24px;margin-top:-28px;position:relative;z-index:2;max-width:600px;margin-left:auto;margin-right:auto}}
.nav-card{{background:white;border-radius:16px;padding:18px 14px 14px;text-align:center;box-shadow:0 2px 12px rgba(0,0,0,0.06);cursor:pointer;transition:transform 0.2s,box-shadow 0.2s,border-color 0.2s;border:2px solid transparent}}
.nav-card:hover{{transform:translateY(-3px);box-shadow:0 6px 24px rgba(0,0,0,0.1)}}.nav-card.active{{border-color:var(--green)}}
.nav-card .nc-score{{font-size:30px;font-weight:300;line-height:1;margin-bottom:2px}}.nav-card .nc-unit{{font-size:13px;color:var(--gray)}}
.nav-card .nc-label{{font-size:11px;font-weight:600;color:var(--gray);margin-top:6px;text-transform:uppercase;letter-spacing:0.5px}}
.nav-card .nc-bar{{height:4px;background:var(--lt-gray);border-radius:2px;margin-top:8px;overflow:hidden}}.nav-card .nc-bar-fill{{height:100%;border-radius:2px}}
.nav-card .nc-cta{{display:flex;align-items:center;justify-content:center;gap:3px;margin-top:8px;font-size:11px;font-weight:500;color:var(--gray);transition:color 0.2s}}
.nav-card:hover .nc-cta{{color:var(--dk-green)}}.nav-card:hover .nc-cta svg{{transform:translateX(2px)}}
.nc-cta svg{{width:12px;height:12px;transition:transform 0.2s}}
.content{{max-width:600px;margin:0 auto;padding:28px 24px 40px}}.panel{{display:none}}.panel.active{{display:block}}
.insight-box{{background:white;border-radius:16px;padding:24px;margin-bottom:20px;box-shadow:0 1px 4px rgba(0,0,0,0.04)}}
.insight-box h3{{font-size:13px;font-weight:600;margin-bottom:10px;text-transform:uppercase;letter-spacing:0.4px}}
.insight-box p{{font-size:15px;line-height:1.7;color:#555}}.hl{{font-weight:600;color:var(--black)}}
.dual-cards{{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:20px}}
.dual-card{{background:white;border-radius:14px;padding:20px;box-shadow:0 1px 4px rgba(0,0,0,0.04)}}
.dual-card h4{{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.4px;margin-bottom:14px}}
.dual-card ul{{list-style:none;padding:0}}.dual-card li{{font-size:14px;line-height:1.5;padding:7px 0;border-bottom:1px solid #f5f3ef}}
.dual-card li:last-child{{border-bottom:none}}.dual-card .sec{{font-size:11px;color:#ccc}}.dual-card .sc{{font-weight:500;float:right}}
.micro-section{{margin-bottom:20px}}.micro-section h3{{font-size:13px;font-weight:600;margin-bottom:14px;text-transform:uppercase;letter-spacing:0.4px}}
.micro-cards{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px}}
.micro-card{{background:white;border-radius:14px;padding:16px;box-shadow:0 1px 4px rgba(0,0,0,0.04);text-align:center}}
.micro-card .mc-name{{font-size:13px;font-weight:500;margin-bottom:6px;line-height:1.3;min-height:34px;display:flex;align-items:flex-end;justify-content:center}}
.micro-card .mc-score{{font-size:22px;font-weight:300;line-height:1;margin-bottom:2px}}.micro-card .mc-raw{{font-size:11px;color:#bbb}}
.micro-card .mc-bar{{height:4px;background:#f0ede8;border-radius:2px;margin-top:10px;overflow:hidden}}.micro-card .mc-bar-fill{{height:100%;border-radius:2px}}
.diff-chart{{background:white;border-radius:16px;padding:24px;margin-bottom:20px;box-shadow:0 1px 4px rgba(0,0,0,0.04)}}
.diff-chart h3{{font-size:13px;font-weight:600;margin-bottom:4px;text-transform:uppercase;letter-spacing:0.4px;color:var(--dk-green)}}
.diff-sub{{font-size:13px;color:#aaa;margin-bottom:20px}}.dsec{{font-size:13px;font-weight:600;margin-bottom:10px}}
.dr{{display:flex;align-items:center;margin-bottom:12px}}.dr:last-child{{margin-bottom:0}}
.dl{{width:60px;font-size:13px;color:#888;font-weight:500;flex-shrink:0}}
.dt{{flex:1;height:26px;background:#f5f3ef;border-radius:8px;overflow:hidden}}
.df{{height:100%;border-radius:8px;display:flex;align-items:center;padding-left:10px;font-size:12px;font-weight:600;color:white;min-width:36px}}
.bg{{background:var(--green)}}.by{{background:var(--orange)}}.br{{background:var(--pink)}}
.ds{{font-size:12px;color:#bbb;margin-left:8px;flex-shrink:0;width:38px;text-align:right}}
.skills-section h3{{font-size:13px;font-weight:600;margin-bottom:14px;text-transform:uppercase;letter-spacing:0.4px}}
.skill-card{{background:white;border-radius:14px;padding:20px;margin-bottom:10px;box-shadow:0 1px 4px rgba(0,0,0,0.04);cursor:pointer;transition:box-shadow 0.2s}}
.skill-card:hover{{box-shadow:0 2px 12px rgba(0,0,0,0.08)}}
.sk-top{{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:12px}}
.sk-name{{font-size:16px;font-weight:500;margin-bottom:4px}}
.sk-tag{{display:inline-block;font-size:11px;font-weight:600;padding:3px 10px;border-radius:20px;letter-spacing:0.2px}}
.tg{{background:#e6faf2;color:var(--dk-green)}}.td{{background:#fff3e0;color:#bf542f}}.tw{{background:#fce4ec;color:#c0476b}}
.sk-pct{{font-size:28px;font-weight:300;line-height:1}}.sk-pct-u{{font-size:14px}}
.sk-bar{{height:6px;background:#f0ede8;border-radius:3px;overflow:hidden}}.sk-bar-fill{{height:100%;border-radius:3px}}
.sk-detail{{max-height:0;overflow:hidden;transition:max-height 0.3s ease}}.skill-card.expanded .sk-detail{{max-height:300px}}
.sk-diff-grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(85px,1fr));gap:8px;margin-top:16px;padding-top:16px;border-top:1px solid #f0ede8}}
.sk-diff-item{{text-align:center;padding:10px 6px;background:#faf8f5;border-radius:10px}}
.sk-diff-item .dn{{font-size:11px;color:#aaa;font-weight:500;margin-bottom:3px}}.sk-diff-item .dv{{font-size:18px;font-weight:500}}.sk-diff-item .dc{{font-size:11px;color:#ccc}}
.back{{display:inline-flex;align-items:center;gap:6px;font-size:14px;font-weight:500;color:var(--dk-green);cursor:pointer;margin-bottom:20px;padding:8px 0}}
.back:hover{{opacity:0.7}}.back svg{{width:16px;height:16px}}
.cta{{background:linear-gradient(135deg,var(--green) 0%,var(--lt-green) 100%);border-radius:16px;padding:32px 24px;margin-top:28px;text-align:center;position:relative;overflow:hidden}}
.cta::before{{content:'';position:absolute;top:-30px;right:-30px;width:120px;height:120px;background:radial-gradient(circle,rgba(255,255,255,0.3) 0%,transparent 70%);border-radius:50%}}
.cta-h{{font-size:18px;font-weight:600;color:var(--black);margin-bottom:8px;position:relative}}.cta-p{{font-size:14px;color:rgba(20,20,20,0.7);line-height:1.6;max-width:400px;margin-left:auto;margin-right:auto;position:relative}}
.footer{{text-align:center;padding:28px 24px}}.footer img{{width:48px;opacity:0.3;margin-bottom:12px}}
.footnote{{font-size:12px;color:#888;max-width:480px;margin:0 auto;line-height:1.7;padding:0 24px}}
@media(max-width:480px){{.hero{{padding:24px 20px 44px}}.student-name{{font-size:26px;letter-spacing:-0.4px}}.student-meta{{margin-bottom:24px}}.score-ring-area{{margin-bottom:12px}}.score-ring{{width:130px;height:130px}}.ring-label .pct{{font-size:32px}}.ring-label .pct-unit{{font-size:15px}}.ring-label .raw{{font-size:11px}}.readiness-scale{{padding:0}}.seg-label-name{{font-size:9px}}.seg-label-sub{{font-size:8px}}.seg-label-range{{font-size:8px}}.seg-marker-label{{font-size:9px;padding:3px 8px}}.score-nav{{gap:6px;padding:0 16px}}.nav-card{{padding:12px 8px 10px;border-radius:12px}}.nav-card .nc-score{{font-size:22px}}.nav-card .nc-unit{{font-size:11px}}.nav-card .nc-label{{font-size:10px;letter-spacing:0.3px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}.nav-card .nc-cta{{font-size:10px}}.dual-cards{{grid-template-columns:1fr}}.micro-cards{{grid-template-columns:1fr}}.content{{padding:24px 16px 40px}}}}
</style>
</head>
<body>
<div id="lock-screen" class="lock-screen">
  <img class="lock-logo" src="data:image/png;base64,{logo_b64}" alt="noon">
  <div class="lock-title">Your GAT Report</div>
  <div class="lock-sub">Enter the last 4 digits of your phone number</div>
  <div class="pin-inputs"><input type="number" maxlength="1" id="p1" autofocus><input type="number" maxlength="1" id="p2"><input type="number" maxlength="1" id="p3"><input type="number" maxlength="1" id="p4"></div>
  <div class="lock-error" id="lock-error"></div>
</div>
<div id="report" class="lock-hidden">
<div class="view active">
  <div class="hero">
    <img class="hero-logo" src="data:image/png;base64,{logo_b64}" alt="noon">
    <div class="student-name">{s['name']}</div>
    <div class="student-meta">Grade {grade_num}  /  GAT Diagnostic  /  {s['date_str']}</div>
    <div class="score-ring-area"><div class="score-ring">
      <svg viewBox="0 0 170 170" width="100%" height="100%"><circle class="ring-bg" cx="85" cy="85" r="73"/><circle class="ring-fill" cx="85" cy="85" r="73"/></svg>
      <div class="ring-label"><div class="pct">{overall_pct}<span class="pct-unit">%</span></div><div class="raw">{overall} / {overall_total}</div></div>
    </div></div>
    {readiness_bar_html(overall_pct)}
  </div>
  <div class="score-nav">
    <div class="nav-card active" onclick="go('overview')"><div class="nc-score" style="color:var(--black)">{overall_pct}<span class="nc-unit">%</span></div><div class="nc-label">Overall</div><div class="nc-bar"><div class="nc-bar-fill" style="width:{overall_pct}%;background:var(--green)"></div></div><div class="nc-cta">Overview</div></div>
    <div class="nav-card" onclick="go('verbal')"><div class="nc-score" style="color:var(--purple)">{verbal_pct}<span class="nc-unit">%</span></div><div class="nc-label">Verbal</div><div class="nc-bar"><div class="nc-bar-fill" style="width:{verbal_pct}%;background:var(--purple)"></div></div><div class="nc-cta">Explore <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 6l6 6-6 6"/></svg></div></div>
    <div class="nav-card" onclick="go('quant')"><div class="nc-score" style="color:var(--blue)">{quant_pct}<span class="nc-unit">%</span></div><div class="nc-label">Quantitative</div><div class="nc-bar"><div class="nc-bar-fill" style="width:{quant_pct}%;background:var(--blue)"></div></div><div class="nc-cta">Explore <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M9 6l6 6-6 6"/></svg></div></div>
  </div>
  <div class="content">
    <div id="p-overview" class="panel active">
      <div class="insight-box"><h3 style="color:var(--dk-green)">Your snapshot</h3><p>{snapshot}</p></div>
      <div class="dual-cards">
        <div class="dual-card"><h4 style="color:var(--dk-green)">Strengths</h4><ul>{dual_list(strengths, '--dk-green')}</ul></div>
        <div class="dual-card"><h4 style="color:var(--pink)">Focus areas</h4><ul>{dual_list(focus, '--pink')}</ul></div>
      </div>
      {"<div class='micro-section'><h3 style='color:var(--dk-green)'>Skills you've nailed</h3><div class='micro-cards'>" + nailed_html + "</div></div>" if nailed_html else ""}
      {"<div class='micro-section'><h3 style='color:var(--pink)'>Top skills to practice</h3><div class='micro-cards'>" + practice_html + "</div></div>" if practice_html else ""}
      <div class="diff-chart">
        <h3>Performance by difficulty</h3><div class="diff-sub">How you performed as questions got harder</div>
        <div style="margin-bottom:20px"><div class="dsec" style="color:var(--purple)">Verbal</div>{diff_bars(v_diff)}</div>
        <div><div class="dsec" style="color:var(--blue)">Quantitative</div>{diff_bars(q_diff)}</div>
      </div>
      <div class="cta"><div class="cta-h">95%+ is closer than you think</div><div class="cta-p">Every student who puts in targeted practice to address knowledge gaps can get there. Your diagnostic mapped exactly where to focus. Noon builds your prep plan around these gaps, so every session counts.</div></div>
    </div>
    <div id="p-verbal" class="panel">
      <div class="back" onclick="go('overview')"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M15 18l-6-6 6-6"/></svg> Back to overview</div>
      <div class="insight-box" style="border-left:3px solid var(--purple);"><h3 style="color:var(--purple)">Verbal: {v_score} / {v_total} ({verbal_pct}%)</h3><p>{verbal_insight}</p></div>
      <div class="skills-section"><h3 style="color:var(--purple)">Skills breakdown</h3>{verbal_cards}</div>
    </div>
    <div id="p-quant" class="panel">
      <div class="back" onclick="go('overview')"><svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M15 18l-6-6 6-6"/></svg> Back to overview</div>
      <div class="insight-box" style="border-left:3px solid var(--blue);"><h3 style="color:var(--blue)">Quantitative: {q_score} / {q_total} ({quant_pct}%)</h3><p>{quant_insight}</p></div>
      <div class="skills-section"><h3 style="color:var(--blue)">Skills breakdown</h3>{quant_cards}</div>
    </div>
  </div>
  <div class="footer"><img src="data:image/png;base64,{logo_b64}" alt="noon"><p class="footnote">This report is not a strict judgement. Its purpose is to provide helpful insights to support your preparation for the test. There are many factors that affect the validity of the report, use discretion in interpreting the results.</p></div>
</div>
</div>
<script>
const PIN='{s["pin"]}';
const inputs=[document.getElementById('p1'),document.getElementById('p2'),document.getElementById('p3'),document.getElementById('p4')];
const lockError=document.getElementById('lock-error');
inputs.forEach((inp,i)=>{{inp.addEventListener('input',function(){{this.value=this.value.slice(-1);if(this.value&&i<3)inputs[i+1].focus();if(i===3&&this.value)checkPin()}});inp.addEventListener('keydown',function(e){{if(e.key==='Backspace'&&!this.value&&i>0)inputs[i-1].focus()}});}});
function checkPin(){{const entered=inputs.map(i=>i.value).join('');if(entered.length===4){{if(entered===PIN){{document.getElementById('lock-screen').style.display='none';document.getElementById('report').classList.remove('lock-hidden')}}else{{lockError.textContent='Incorrect code, please try again';inputs.forEach(i=>{{i.classList.add('error');i.value=''}});inputs[0].focus();setTimeout(()=>{{inputs.forEach(i=>i.classList.remove('error'));lockError.textContent=''}},1500)}}}}}}
function go(name){{document.querySelectorAll('.panel').forEach(p=>p.classList.remove('active'));document.getElementById('p-'+name).classList.add('active');document.querySelectorAll('.nav-card').forEach((c,i)=>{{c.classList.remove('active');if((name==='overview'&&i===0)||(name==='verbal'&&i===1)||(name==='quant'&&i===2))c.classList.add('active')}});window.scrollTo({{top:document.querySelector('.score-nav').offsetTop-20,behavior:'smooth'}})}}
</script>
</body>
</html>'''

    return html


def slugify(name):
    return re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Generate GAT diagnostic reports')
    parser.add_argument('--question-bank', required=True, help='Path to GAT_Question_Bank.xlsx')
    parser.add_argument('--responses', required=True, help='Path to response CSV')
    parser.add_argument('--output-dir', default='docs/', help='Output directory for HTML files')
    parser.add_argument('--logo-path', required=True, help='Path to noon logo PNG')
    args = parser.parse_args()

    # Load logo
    with open(args.logo_path, 'rb') as f:
        logo_b64 = base64.b64encode(f.read()).decode()

    # Load data
    verbal_q, quant_q = load_question_bank(args.question_bank)
    students = load_responses(args.responses, verbal_q, quant_q)

    os.makedirs(args.output_dir, exist_ok=True)

    print(f"Generating reports for {len(students)} students...")

    manifest = []
    for student in students:
        slug = slugify(student['name'])
        filename = f"{slug}.html"
        filepath = os.path.join(args.output_dir, filename)

        html = generate_report(student, logo_b64)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html)

        manifest.append({
            'name': student['name'],
            'slug': slug,
            'filename': filename,
            'pin': student['pin'],
        })
        print(f"  ✓ {student['name']} → {filename} (PIN: {student['pin']})")

    # Save manifest (for your reference, not public)
    with open(os.path.join(args.output_dir, '_manifest.json'), 'w') as f:
        json.dump(manifest, f, indent=2)

    print(f"\nDone. {len(manifest)} reports in {args.output_dir}")
    print(f"Manifest saved to {args.output_dir}/_manifest.json")


if __name__ == '__main__':
    main()
